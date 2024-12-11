import os
from ExpInfo import *
from typing import List
from openpyxl import load_workbook
import xlsxwriter
import numpy as np
import logging
logger=logging.getLogger("LogWindow")
import PathManager
from scipy.stats import pearsonr, spearmanr

def PostProcess(exp_setting:ExpSetting,output_folder):
    if exp_setting.save_format == SaveFormat.CSV:
        post_fix = 'csv'
    else:
        post_fix = "xlsx"
    
    output_file=os.path.join(output_folder,"outlier_result.%s" %post_fix)
    all_files = os.listdir(output_folder)
    all_files = [x for x in all_files if x.endswith(post_fix)]
    
    if len(all_files) == 0:
        logger.error("No files found!")
        return
    
    all_subjects_scores={}
    all_subjects=[]
    
    for file_name in all_files:
        if file_name in  output_file:
            continue
        subject_name = file_name.split('.')[0]
        subject_scores = GetScore(os.path.join(output_folder,file_name))
        all_subjects_scores[subject_name]=subject_scores
        all_subjects.append(subject_name)
    
    all_score_names=list(subject_scores.keys())

    type_scores={}
    # first get all the results into matrix
    for score_name in all_score_names:
        type_scores[score_name]=[]
        for subject_name in all_subjects:
            type_scores[score_name].append(all_subjects_scores[subject_name][score_name])
    
        type_scores[score_name]=np.vstack(type_scores[score_name])
    
    # then calculate the plcc for each subjects
    all_subjects_plcc={}
    for subject_index,subject_name in enumerate(all_subjects):
        all_subjects_plcc[subject_name]=[]
        for score_name in all_score_names:
            all_subjects_plcc[subject_name].append(CalSingleSROCC(type_scores[score_name],subject_index))
    
    # save the plcc
    if post_fix == 'csv':
        SaveCSVPLCC(all_subjects_plcc,all_score_names,output_file)
    else:
        SaveExcelPLCC(all_subjects_plcc,all_score_names,output_file)

def ExportMOS(all_persons: List[PersonInfo]):
    all_subject_names=[]
    all_subject_files=[]

    for person in all_persons:
        all_subject_names.append(person.name)
        all_subject_files.append(person.result_file)
    
    all_subject_scores=GetAllSubjectScores(all_subject_files)
    all_score_names=list(all_subject_scores.keys())

    ret_mean_score={}
    for score_name in all_score_names:
        ret_mean_score[score_name]=all_subject_scores[score_name].mean(0)
        ret_mean_score[score_name]=ret_mean_score[score_name].tolist()
    
    return ret_mean_score

def ExportSROCC(all_persons: List[PersonInfo]):
    all_subject_names=[]
    all_subject_files=[]

    for person in all_persons:
        all_subject_names.append(person.name)
        all_subject_files.append(person.result_file)
    
    all_subject_scores=GetAllSubjectScores(all_subject_files)
    all_score_names=list(all_subject_scores.keys())

    ret_srocc={}
    for score_name in all_score_names:
        ret_srocc[score_name]=ScipyCalSROCC(all_subject_scores[score_name])
    
    return ret_srocc

def ExportPLCC(all_persons: List[PersonInfo]):
    all_subject_names=[]
    all_subject_files=[]

    for person in all_persons:
        all_subject_names.append(person.name)
        all_subject_files.append(person.result_file)
    
    all_subject_scores=GetAllSubjectScores(all_subject_files)
    all_score_names=list(all_subject_scores.keys())

    ret_plcc={}
    for score_name in all_score_names:
        ret_plcc[score_name]=ScipyCalPLCC(all_subject_scores[score_name])
    
    return ret_plcc

def ExportAll(all_persons: List[PersonInfo]):
    all_subject_names=[]
    all_subject_files=[]

    for person in all_persons:
        all_subject_names.append(person.name)
        all_subject_files.append(person.result_file)
    
    all_subject_scores=GetAllSubjectScores(all_subject_files)
    all_score_names=list(all_subject_scores.keys())

    ret_srocc={}
    ret_plcc={}
    for score_name in all_score_names:
        ret_srocc[score_name]=ScipyCalSROCC(all_subject_scores[score_name])
        ret_plcc[score_name]=ScipyCalPLCC(all_subject_scores[score_name])
    
    return ret_srocc,ret_plcc

def ScipyCalSROCC(score_matrix):
    # score_matrix: subject_num x img_num
    subject_num, img_num=score_matrix.shape

    ret_srocc=[0 for i in range(subject_num)]
    mean_score=score_matrix.mean(0)    

    for i in range(subject_num):
        cur_srocc,_=spearmanr(score_matrix[i,:],mean_score)
        ret_srocc[i]=cur_srocc

    return ret_srocc

def ScipyCalPLCC(score_matrix):
    # score_matrix: subject_num x img_num
    subject_num, img_num=score_matrix.shape

    ret_plcc=[0 for i in range(subject_num)]
    mean_score=score_matrix.mean(0)    

    for i in range(subject_num):
        cur_plcc,_=pearsonr(score_matrix[i,:],mean_score)
        ret_plcc[i]=cur_plcc

    return ret_plcc


'''
def CalSingleSROCC(score_matrix,subject_index):
    score_1=score_matrix[subject_index,:]
    score_2=score_matrix.sum(0)-score_1
    score_2=score_2/(score_matrix.shape[0]-1)

    plcc=np.corrcoef(score_1,score_2)
    plcc=plcc[0,1]
    return plcc
'''
    
def SaveExcelPLCC(all_subjects_plcc,all_score_names,save_file):
    workbook=xlsxwriter.Workbook(save_file)
    work_sheet=workbook.add_worksheet('PLCC')
    work_sheet.write(0,0,'Subject Name')
    for i,score_name in enumerate(all_score_names):
        work_sheet.write(0,i+1,score_name)

    all_subject_names=list(all_subjects_plcc.keys())
    for i,subject_name in enumerate(all_subject_names):
        work_sheet.write(i+1,0,subject_name)
        for j in range(len(all_score_names)):
            work_sheet.write(i+1,j+1,all_subjects_plcc[subject_name][j])
    workbook.close()

def SaveCSVPLCC(all_subjects_plcc,all_score_names,save_file):
    with open(save_file,'w') as fid:
        fid.write('Subject Name')

        for score_name in all_score_names:
            fid.write(',%s' %score_name)
        fid.write('\n')

        all_subject_names=list(all_subjects_plcc.keys())
        for i,subject_name in enumerate(all_subject_names):
            fid.write('%s' %subject_name)
            for j in range(len(all_score_names)):
                fid.write(',%f' %all_subjects_plcc[subject_name][j])
            fid.write('\n') 

    
def ReadSubjectScore(file_name):
    if not os.path.exists(file_name):
        logger.error(f"The file {file_name} does not exist!")
        return None
    all_content=None
    if file_name.endswith(".csv"):
        all_content=ReadCSVAllData(file_name)
    else:
        all_content=ReadExcelAllData(file_name)
    
    img_num=len(all_content)-1
    
    title_line=all_content[0]
    all_score_names=title_line[2:]

    ret_score={}
    ret_file_name=[None for i in range(img_num)]
    for score_name in all_score_names:
        ret_score[score_name]=[0 for i in range(img_num)]

    for line in all_content[1:]:
        img_index = int(line[0])
        ret_file_name[img_index]=line[1]
        cur_socres=line[2:]
        for i in range(len(cur_socres)):
            ret_score[all_score_names[i]][img_index]=int(cur_socres[i])

    return ret_score,ret_file_name

def GetAllSubjectScores(all_files):
    all_subjects_scores={}
    for file_name in all_files:
        cur_score,cur_file_name=ReadSubjectScore(file_name)
        for score_name in cur_score.keys():
            if score_name not in all_subjects_scores.keys():
                all_subjects_scores[score_name]=[]
            all_subjects_scores[score_name].append(cur_score[score_name])
    
    for score_name in all_subjects_scores.keys():
        all_subjects_scores[score_name]=np.array(all_subjects_scores[score_name])
    return all_subjects_scores

def ReadExcelAllData(file_name):
    wb=load_workbook(file_name)
    sheets=wb.worksheets
    sheet=sheets[0]
    all_data=[]
    for row in sheet:
        row_data=[x.value for x in row]
        all_data.append(row_data)
    return all_data

def ReadCSVAllData(file_name):
    with open(file_name,'r') as fid:
        lines=fid.readlines()
    all_data=[]
    for line in lines:
        line_data=line.split(',')
        all_data.append(line_data)
    return all_data
    

    
    
    
    
    
    
    
    
    
    
    
    
    
    
