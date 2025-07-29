#########################################################
## To Manage the path of the project
#########################################################

import os
import json
import sys
sys.path.append('../Widgets')
import logging
logger=logging.getLogger("LogWindow")
import openpyxl

software_version='3.0'

inner_views_refocusing_path="views_refocusing"
inner_show_views_path="show_views"
inner_show_refocusing_path="show_refocusing"
compair_folder="PairComparison"
training_pair_comparison_root="training"
test_pair_comparison_root="test"
passive_view_video_name="view"
passive_refocusing_video_name="refocus"
inner_depth_map="depth.png"
project_intermediate_data="IntermediateData"
project_training_data="TrainingData"
project_test_data="TestData"
thumbnail_name="thumbnail"
refocusing_mask_key_word="mask"
active_refocusing_folder_key_word="refocus"
ffmpeg_path="ffmpeg.exe"
btn_height=50
btn_width=226

scoring_btn_height=40
scoring_btn_width=229
table_spaceing=25
hint_text_font_size=60
scoring_table_point_size=20
scoring_table_width=290
scoring_group_box_width=250
table_horizontal_gap=20
talbe_vertical_gap=10

#ffmpeg_path="D:/ffmpeg/bin/ffmpeg.exe"
# lambda file will be deprecated
lambda_file="lambda.txt"

# output folders
subject_results_folder="SubjectResults"
all_subject_info_file='all_subject_info.csv'
mos_file='MOS.csv'
srocc_file='SROCC.csv'
plcc_file='PLCC.csv'
all_results_file='AllResults.xlsx'
dscs_folder="DSCS_PC"

# cache folders
cache_folder="Cache"
training_cache_folder="Training"
testing_cache_folder="Testing"
cache_thumbnail='thumbnail.png'
cache_desc='LFIMapping'

def VersionCMP(v1,v2):
    v1=v1.split('.')
    v2=v2.split('.')

    loop_max=max(len(v1),len(v2))
    if loop_max>len(v1):
        v1+=[0]*(loop_max-len(v1))
    if loop_max>len(v2):
        v2+=[0]*(loop_max-len(v2))
    for i in loop_max:
        if int(v1[i])>int(v2[i]):
            return 1
        elif int(v1[i])<int(v2[i]):
            return -1
    return 0

def OpenPath(path):
    if os.path.isfile(path):
        path=os.path.dirname(path)
    if os.name == 'nt':
        os.startfile(path)
    else:
        opener = 'open' if sys.platform == 'darwin' else 'xdg-open'
        os.system([opener, path])

'''
Maybe a class to derive and manage the IO paths?
Automatically generate the path during the preprocessing.
Regarding the custom and manual preprocessing, init the paths during the configuration stage.
'''

class ExpShowPathManager:
    def __init__(self,in_path,mode="training",use_pair_comparison=False,video_post_fix='mp4',out_img_post_fix='png',pair_comparison_index="0"):
        '''
        in_path is the path of a single LFI
        the pair comparison index may be int or str
        '''
        self.root_path=in_path
        self.video_post_fix=video_post_fix
        self.out_img_post_fix=out_img_post_fix
        self.mode=mode
        self.pair_comparison=use_pair_comparison
        self.cmp_index=pair_comparison_index

        self.refocusing_views_path=os.path.join(in_path,inner_views_refocusing_path)

        self.show_view_path=os.path.join(in_path,inner_show_views_path)
        self.show_refocusing_path=os.path.join(in_path,inner_show_refocusing_path)

        self.passive_view_video=os.path.join(in_path,inner_show_views_path,passive_view_video_name)
        self.passive_refocusing_video=os.path.join(in_path,inner_show_refocusing_path,passive_refocusing_video_name)

        self.training_comparison_root=os.path.join(in_path,training_pair_comparison_root,f'{compair_folder}_{pair_comparison_index}')
        self.test_comparison_root=os.path.join(in_path,test_pair_comparison_root,f'{compair_folder}_{pair_comparison_index}')

        self.training_comparison_show_view_path=os.path.join(self.training_comparison_root,inner_show_views_path)
        self.training_comparison_passivive_view_video=os.path.join(self.training_comparison_show_view_path,passive_view_video_name)
        self.training_comparison_show_refocusing_path=os.path.join(self.training_comparison_root,inner_show_refocusing_path)
        self.training_comparison_passive_refocusing_video=os.path.join(self.training_comparison_show_refocusing_path,passive_refocusing_video_name)

        self.test_comparison_show_view_path=os.path.join(self.test_comparison_root,inner_show_views_path)
        self.test_comparison_passivive_view_video=os.path.join(self.test_comparison_show_view_path,passive_view_video_name)
        self.test_comparison_show_refocusing_path=os.path.join(self.test_comparison_root,inner_show_refocusing_path)
        self.test_comparison_passive_refocusing_video=os.path.join(self.test_comparison_show_refocusing_path,passive_refocusing_video_name)
    
        if not self.pair_comparison:
            self._show_view_path=self.show_view_path
            self._show_refocus_path=self.show_refocusing_path
            self._passive_view_video=self.passive_view_video
            self._passive_refocus_video=self.passive_refocusing_video
        else:
            if self.mode ==  "training":
                self._show_view_path=self.training_comparison_show_view_path
                self._show_refocus_path=self.training_comparison_show_refocusing_path
                self._passive_view_video=self.training_comparison_passivive_view_video
                self._passive_refocus_video=self.training_comparison_passive_refocusing_video
            else:
                self._show_view_path=self.test_comparison_show_view_path
                self._show_refocus_path=self.test_comparison_show_refocusing_path
                self._passive_view_video=self.test_comparison_passivive_view_video
                self._passive_refocus_video=self.test_comparison_passive_refocusing_video
        
        self.depth_map_path=os.path.join(self._show_refocus_path,inner_depth_map)
    
    def Get_refocusing_view_path(self):
        return self.refocusing_views_path

    def Get_depth_map_path(self):
        return self.depth_map_path

    def Get_show_view_path(self):
        return self._show_view_path
    def Get_show_refocus_path(self):
        return self._show_refocus_path

    def Get_passive_view_video_name(self):
        return self._passive_view_video
    def Get_passive_view_video_path(self):
        return f"{self._passive_view_video}.{self.video_post_fix}"
    
    def Get_passive_refocus_video_name(self):
        return self._passive_refocus_video
    def Get_passive_refocus_video_path(self):
        return f"{self._passive_refocus_video}.{self.video_post_fix}"
    

class SoftWarePathManager():
    def __init__(self):
        # TODO: set software description with input json file
        self._software_path=''
        self._software_version=''
        self._logs_path='./Logs'
        self._log_level="INFO"
        self.config_keys=[
            "Software_Path",
            "Software_Version",
            "Logs_Path",
            "Log_Level"
        ]
        self.log_level_dict={
            'INFO':logging.INFO,
            "DEBUG":logging.DEBUG,
            "WARNING":logging.WARNING,
            "ERROR": logging.ERROR
        }
        self.config={}
        self.config['Software_Path']=self._software_path
        self.config['Software_Version']=self._software_version
        self.config['Logs_Path']=self._logs_path
        self.config['Log_Level']=self._log_level

    def CheckInnerPath(self,path):
        if not os.path.exists(path):
            os.makedirs(path)
    
    def SaveInfo(self):
        with open(self.file_path,'w') as fid:
            json.dump(self.config,fid,indent=4)
        
    @property
    def software_path(self):
        return self._software_path
    
    @software_path.setter
    def software_path(self,value):
        self._software_path=value
        self.config['Software_Path']=value

    @property
    def software_version(self):
        return self._software_version
    
    @software_version.setter
    def software_version(self,value):
        self._software_version=value
        self.config["Software_Version"]=value
    
    @property
    def logs_path(self):
        return self._logs_path
    @logs_path.setter
    def logs_path(self,value):
        self._logs_path=value
        self.config['Logs_Path']=value

    @property
    def log_level(self):
        return self._log_level
    @log_level.setter
    def log_level(self,value):
        if isinstance(value,str):
            self._log_level=self.log_level_dict[value]
            self.config['Log_Level']=value
        else:
            self._log_level=value
            for key,dict_value in self.log_level_dict.items():
                if value==dict_value:
                    self.config['Log_Level']=key
                    break
    
    @staticmethod
    def CheckInitFile(input_json):
        config_keys=[
            "Software_Path",
            "Software_Version",
            "Logs_Path",
            "Log_Level"
        ]
        with open(input_json,'r') as fid:
            cur_config=json.load(fid)
        for config_key in config_keys:
            if config_key not in cur_config.keys():
                return False
        return True
    
    @staticmethod
    def ReadLogLevelOnly(input_json):
        if not os.path.exists(input_json):
            return None
        with open(input_json,'r') as fid:
            cur_config=json.load(fid)
        if 'Log_Level' not in cur_config.keys():
            return None
        log_level=cur_config['Log_Level']
        if log_level.upper() not in ['INFO',"ERROR","WARNING","DEBUG"]:
            return None
        return log_level

def GetSubjectResultFolder(project_path):
    return os.path.join(project_path,subject_results_folder)

def GetAllSubjectInfoFile(project_path):
    return os.path.join(project_path,subject_results_folder,all_subject_info_file)

def GetSubjectResultFile(project_path,subject_file_name):
    return os.path.join(project_path,subject_results_folder,subject_file_name)

def _open_path(path):
    if not os.path.exists(path):
        logger.warning(f"The path {path} does not exist!")
        return False
    else:
        if sys.platform == "win32":
            os.system(f'explorer {path}')
        elif sys.platform == "linux":
            os.system(f'xdg-open {path}')
        else:
            os.system(f'open {path}')
        return True        

def OpenPath(path):
    if os.path.isfile(path):
        path=os.path.dirname(path)
    _open_path(path)

def DeleteFile(path):
    if os.path.exists(path):
        os.remove(path)

def ReadSubjectResult_CSV(file_path):
    if not os.path.exists(file_path):
        logger.error(f"The file {file_path} does not exist!")
        return None
    all_content=[]
    with open(file_path,'r') as fid:
        for line in fid.readlines():
            all_content.append(line.strip().split(','))
    return all_content

def ReadSubjectResult_Excel(file_path):
    #actually, only 1 sheet for each
    if not os.path.exists(file_path):
        logger.error(f"The file {file_path} does not exist!")
        return None
    work_book=openpyxl.load_workbook(file_path)
    subject_sheet=work_book.worksheets[0]
    all_content=[]
    for row in subject_sheet.rows:
        all_content.append([cell.value for cell in row])
    work_book.close()
    return all_content

def ReadSubjectResult(project_file):
    if not os.path.exists(project_file):
        logger.error(f"The file {project_file} does not exist!")
        return None
    if '.csv' in project_file:
        return ReadSubjectResult_CSV(project_file)
    elif '.xlsx' in project_file:
        return ReadSubjectResult_Excel(project_file)
    else:
        logger.error(f"The file {project_file} is not supported!")
        return None

# Do we need to save a backup file if the file exists?
def SaveToCSV(file_path,all_content):
    with open(file_path,'w') as fid:
        for line in all_content:
            line=[str(item) for item in line]
            fid.write(','.join(line))
            fid.write('\n')

def SaveToExcel(file_path,all_content,sheet_name=None):
    # Terrible lib
    # here first check if the file exists,
    # if not, create a new one
    # if yes, make a blank one (remove default sheets), and then write
    if os.path.exists(file_path):
        workbook=openpyxl.load_workbook(file_path)
    else:
        workbook=openpyxl.Workbook()
        for sheet in workbook.worksheets:
            workbook.remove(sheet)
    
    if sheet_name is None:
        sheet_name=os.path.basename(file_path).split('.')[0]
    work_sheet=workbook.create_sheet(sheet_name)
    
    for row_index,line in enumerate(all_content,1):
        for col_index, in_v in enumerate(line,1):
            work_sheet.cell(row=row_index,column=col_index).value=in_v
    
    workbook.save(file_path)
    workbook.close()

def SaveToFile(file_path,content):
    # save to csv or excel
    if '.csv' in file_path:
        SaveToCSV(file_path,content)
    elif '.xlsx' in file_path:
        SaveToExcel(file_path,content)
    else:
        logger.error(f"The file {file_path} is not supported!")

def GetFileNameFromPath(file_path):
    return os.path.basename(file_path)

def GetFileNameWithoutExtension(file_path):
    file_name=GetFileNameFromPath(file_path)
    file_name_without_extension=file_name.split('.')[0]
    return file_name_without_extension

def GetExtension(file_path):
    return os.path.basename(file_path).split('.')[-1]

def GetMOSFileName(file_folder):
    return os.path.join(file_folder,mos_file)

def GetSROCCFileName(file_folder):
    return os.path.join(file_folder,srocc_file)

def GetPLCCFileName(file_folder):
    return os.path.join(file_folder,plcc_file)

def GetAllResultFileName(file_folder):
    return os.path.join(file_folder,all_results_file)

if __name__ == "__main__":
    all_content=[
        ['img_index','img_name','mos'],
        [1,'img1',5],
        [2,'img2',6],
        [0,'img3',7]
    ]

    SaveToCSV('test.csv',all_content)

    SaveToExcel('test.xlsx','test',all_content)